"""
Pruebas de los tres indicadores de la tesis y de su exportación.

Lo que se fija aquí son las decisiones de cálculo que se pueden romper sin que
nada falle a la vista: que las tasas se midan solo sobre lo etiquetado, que un
período sin fraudes confirmados no invente un cero, y —la más fácil de perder—
que el total del rango salga de sumar los casos y dividir una vez, y no de
promediar los porcentajes de cada período.
"""

from datetime import datetime, timedelta

import pytest
from openpyxl import load_workbook
from io import BytesIO

from app.models.fraud_log import FraudLog
from app.models.order import Order, OrderStatus
from app.models.user import UserRole
from app.services import fraud_metrics_service
from app.services.fraud_metrics_service import ZONA_DE_LA_TIENDA

from tests.conftest import cabeceras_de, crear_usuario


async def _evaluacion(
    sesion,
    usuario,
    decision: str,
    cuando: datetime,
    *,
    etiquetada: bool = False,
    fue_fraude: bool = False,
    milisegundos: float = 5.0,
    monto: float = 500.0,
):
    """Un pedido evaluado, opcionalmente ya revisado por un administrador."""
    orden = Order(
        user_id=usuario.id,
        total_amount=monto,
        status={
            "BLOCKED": OrderStatus.REJECTED,
            "REVIEW": OrderStatus.FRAUD_REVIEW,
        }.get(decision, OrderStatus.PENDING),
        shipping_city="Trujillo",
        created_at=cuando,
    )
    sesion.add(orden)
    await sesion.flush()

    sesion.add(
        FraudLog(
            order_id=orden.id,
            fraud_score=0.5,
            decision=decision,
            evaluated_at=cuando,
            detection_time_ms=milisegundos,
            reviewed_at=cuando if etiquetada else None,
            is_actual_fraud=fue_fraude,
        )
    )
    await sesion.commit()
    return orden


@pytest.mark.asyncio
async def test_la_tasa_de_deteccion_solo_cuenta_lo_etiquetado(sesion):
    usuario = await crear_usuario(sesion)
    hoy = datetime.now(ZONA_DE_LA_TIENDA)

    # Dos fraudes confirmados: uno bloqueado (detectado) y otro aprobado (no).
    await _evaluacion(sesion, usuario, "BLOCKED", hoy, etiquetada=True, fue_fraude=True)
    await _evaluacion(sesion, usuario, "APPROVED", hoy, etiquetada=True, fue_fraude=True)
    # Una compra legítima confirmada: no entra en el denominador de la tasa.
    await _evaluacion(sesion, usuario, "APPROVED", hoy, etiquetada=True, fue_fraude=False)
    # Y una sin revisar: tampoco, porque no se sabe qué era.
    await _evaluacion(sesion, usuario, "APPROVED", hoy)

    periodo = (await fraud_metrics_service.historial(sesion, "day", periodos=1))[0]

    assert periodo.evaluaciones == 4
    assert periodo.revisados == 3
    assert periodo.fraudes_reales == 2
    assert periodo.fraudes_detectados == 1
    assert periodo.fraudes_no_detectados == 1
    assert periodo.tasa_de_deteccion == 0.5
    assert periodo.tasa_de_no_deteccion == 0.5


@pytest.mark.asyncio
async def test_una_revision_manual_tambien_cuenta_como_detectado(sesion):
    """
    Un fraude mandado a revisión está detectado aunque no se haya bloqueado:
    lo que importa es que la compra no siguió su curso hasta el cobro.
    """
    usuario = await crear_usuario(sesion)
    hoy = datetime.now(ZONA_DE_LA_TIENDA)
    await _evaluacion(sesion, usuario, "REVIEW", hoy, etiquetada=True, fue_fraude=True)

    periodo = (await fraud_metrics_service.historial(sesion, "day", periodos=1))[0]
    assert periodo.fraudes_detectados == 1
    assert periodo.tasa_de_deteccion == 1.0


@pytest.mark.asyncio
async def test_sin_fraudes_confirmados_la_tasa_no_es_cero_sino_desconocida(sesion):
    """
    Un período con compras legítimas y ningún fraude no tiene tasa de
    detección. Devolver 0 % diría "no detectamos nada", que es una acusación
    falsa contra el modelo; lo cierto es que no hay con qué medirlo.
    """
    usuario = await crear_usuario(sesion)
    hoy = datetime.now(ZONA_DE_LA_TIENDA)
    await _evaluacion(sesion, usuario, "APPROVED", hoy, etiquetada=True, fue_fraude=False)

    periodo = (await fraud_metrics_service.historial(sesion, "day", periodos=1))[0]

    assert periodo.fraudes_reales == 0
    assert periodo.tasa_de_deteccion is None
    assert periodo.tasa_de_no_deteccion is None
    # El tiempo sí se mide: no necesita etiquetas.
    assert periodo.tiempo_medio_ms > 0


@pytest.mark.asyncio
async def test_el_total_del_rango_no_promedia_porcentajes(cliente, sesion):
    """
    El error clásico de este tipo de reporte.

    Un mes con un fraude detectado (100 %) y otro con diez de los que se
    detectaron dos (20 %) dan, promediando porcentajes, un 60 % que no le
    corresponde a nadie. Agrupando los casos son 3 de 11: 27 %.
    """
    await crear_usuario(sesion, email="admin@ejemplo.com", rol=UserRole.ADMIN)
    usuario = await crear_usuario(sesion)

    hoy = datetime.now(ZONA_DE_LA_TIENDA)
    ayer = hoy - timedelta(days=1)

    await _evaluacion(sesion, usuario, "BLOCKED", ayer, etiquetada=True, fue_fraude=True)
    for i in range(10):
        await _evaluacion(
            sesion,
            usuario,
            "BLOCKED" if i < 2 else "APPROVED",
            hoy,
            etiquetada=True,
            fue_fraude=True,
        )

    cabeceras = await cabeceras_de(cliente, "admin@ejemplo.com")
    datos = (
        await cliente.get(
            "/api/v1/fraud/history?granularity=day&periods=2", headers=cabeceras
        )
    ).json()

    assert datos["total_actual_frauds"] == 11
    assert datos["total_detected_frauds"] == 3
    assert datos["detection_rate"] == pytest.approx(3 / 11, abs=1e-4)
    # El promedio de los dos porcentajes sería 0.6: si sale eso, se rompió.
    assert datos["detection_rate"] < 0.5


@pytest.mark.asyncio
async def test_el_reporte_en_excel_trae_los_mismos_numeros_que_el_panel(cliente, sesion):
    await crear_usuario(sesion, email="admin@ejemplo.com", rol=UserRole.ADMIN)
    usuario = await crear_usuario(sesion)
    hoy = datetime.now(ZONA_DE_LA_TIENDA)

    await _evaluacion(sesion, usuario, "BLOCKED", hoy, etiquetada=True, fue_fraude=True, monto=900.0)
    await _evaluacion(sesion, usuario, "APPROVED", hoy, etiquetada=True, fue_fraude=False, monto=100.0)

    cabeceras = await cabeceras_de(cliente, "admin@ejemplo.com")

    panel = (
        await cliente.get("/api/v1/fraud/history?granularity=day&periods=1", headers=cabeceras)
    ).json()

    respuesta = await cliente.get(
        "/api/v1/fraud/report.xlsx?granularity=day&periods=1", headers=cabeceras
    )
    assert respuesta.status_code == 200
    assert "spreadsheetml" in respuesta.headers["content-type"]
    assert ".xlsx" in respuesta.headers["content-disposition"]

    libro = load_workbook(BytesIO(respuesta.content))
    assert libro.sheetnames == ["Resumen", "Indicadores por período", "Montos"]

    hoja = libro["Indicadores por período"]
    fila = {c.value: hoja.cell(row=2, column=c.column).value for c in hoja[1]}

    assert fila["Evaluadas"] == panel["total_evaluations"] == 2
    assert fila["Fraudes"] == panel["total_actual_frauds"] == 1
    assert fila["Detectados"] == panel["total_detected_frauds"] == 1
    # La tasa viaja como número, no como el texto "100 %": un texto no se
    # promedia ni se grafica en Excel.
    assert fila["Tasa detectados"] == 1.0
    assert isinstance(fila["Tasa detectados"], (int, float))


@pytest.mark.asyncio
async def test_el_reporte_es_solo_para_administradores(cliente, sesion):
    await crear_usuario(sesion, email="cliente@ejemplo.com")
    cabeceras = await cabeceras_de(cliente, "cliente@ejemplo.com")

    respuesta = await cliente.get("/api/v1/fraud/report.xlsx", headers=cabeceras)
    assert respuesta.status_code == 403
