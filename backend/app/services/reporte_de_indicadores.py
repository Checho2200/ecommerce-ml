"""
El reporte de indicadores del antifraude, en un archivo de Excel.

Existe porque una tesis se sustenta con números que alguien puede abrir,
ordenar y volver a sumar, no con capturas de pantalla. El archivo trae las tres
hojas que hacen falta para eso: qué mide cada indicador, la serie completa por
período, y el total de la ventana.

Los números no se recalculan aquí. Llegan ya armados desde el mismo objeto que
responde `/fraud/history`, así que el archivo no puede desviarse de lo que
enseña el panel: un reporte que no cuadra con la pantalla de la que sale es
peor que no tener reporte.
"""

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.schemas.fraud import FraudHistoryResponse

# Cómo se llama cada escala en la portada del archivo.
ESCALAS = {
    "day": ("Diario", "días"),
    "week": ("Semanal", "semanas"),
    "month": ("Mensual", "meses"),
    "year": ("Anual", "años"),
}

# Los tres indicadores, con la dirección en la que conviene que se muevan. Es
# lo que convierte una tabla de porcentajes en un cuadro de mando: sin decir
# hacia dónde es mejor, un 12 % no significa nada.
INDICADORES = [
    (
        "Tasa de fraudes detectados",
        "De los fraudes confirmados, qué proporción frenó el modelo antes de cobrar.",
        "Debe subir",
    ),
    (
        "Tasa de fraude no detectado",
        "De los fraudes confirmados, qué proporción se aprobó igual y terminó en pérdida.",
        "Debe bajar",
    ),
    (
        "Tiempo de detección",
        "Cuánto tarda el modelo en evaluar una compra, en milisegundos.",
        "Debe bajar",
    ),
]

_AZUL = "0C3A6E"
_BORDE = Border(bottom=Side(style="thin", color="D0D7E2"))


def _encabezado(hoja, fila: int, titulos: list[str]) -> None:
    for columna, titulo in enumerate(titulos, start=1):
        celda = hoja.cell(row=fila, column=columna, value=titulo)
        celda.font = Font(bold=True, color="FFFFFF", size=10)
        celda.fill = PatternFill("solid", fgColor=_AZUL)
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _anchos(hoja, anchos: list[int]) -> None:
    for i, ancho in enumerate(anchos, start=1):
        hoja.column_dimensions[get_column_letter(i)].width = ancho


def _hoja_portada(libro: Workbook, datos: FraudHistoryResponse) -> None:
    hoja = libro.active
    hoja.title = "Resumen"
    _anchos(hoja, [34, 62, 16, 18])

    hoja["A1"] = "Indicadores del sistema antifraude"
    hoja["A1"].font = Font(bold=True, size=15, color=_AZUL)
    hoja["A2"] = "Grupo STS SAC — tienda en línea"
    hoja["A2"].font = Font(size=10, color="5A6878")

    escala, unidad = ESCALAS.get(datos.granularity, ("—", "períodos"))
    hoja["A4"] = "Escala del reporte"
    hoja["B4"] = escala
    hoja["A5"] = "Períodos incluidos"
    hoja["B5"] = f"{len(datos.periods)} {unidad}"
    hoja["A6"] = "Generado el"
    hoja["B6"] = datetime.now().strftime("%d/%m/%Y %H:%M")
    for fila in range(4, 7):
        hoja.cell(row=fila, column=1).font = Font(bold=True, size=10)

    # ── Los tres indicadores ────────────────────────────────────────────────
    _encabezado(hoja, 8, ["Indicador", "Qué mide", "Valor", "Dirección deseada"])

    tasa_deteccion = (
        f"{datos.detection_rate * 100:.1f} %" if datos.detection_rate is not None else "sin datos"
    )
    tasa_no_deteccion = (
        f"{datos.undetected_rate * 100:.1f} %" if datos.undetected_rate is not None else "sin datos"
    )
    valores = [tasa_deteccion, tasa_no_deteccion, f"{datos.average_detection_time_ms:.1f} ms"]

    for i, ((nombre, descripcion, direccion), valor) in enumerate(zip(INDICADORES, valores)):
        fila = 9 + i
        hoja.cell(row=fila, column=1, value=nombre).font = Font(bold=True, size=10)
        hoja.cell(row=fila, column=2, value=descripcion).alignment = Alignment(wrap_text=True)
        celda = hoja.cell(row=fila, column=3, value=valor)
        celda.font = Font(bold=True, size=12, color=_AZUL)
        celda.alignment = Alignment(horizontal="center")
        hoja.cell(row=fila, column=4, value=direccion).alignment = Alignment(horizontal="center")
        for columna in range(1, 5):
            hoja.cell(row=fila, column=columna).border = _BORDE
        hoja.row_dimensions[fila].height = 30

    # ── Sobre qué se midieron ───────────────────────────────────────────────
    hoja["A14"] = "Base de cálculo"
    hoja["A14"].font = Font(bold=True, size=11, color=_AZUL)

    base = [
        ("Compras evaluadas", datos.total_evaluations),
        ("Pasaron al cobro", datos.total_approved),
        ("Retenidas (revisión o bloqueo)", datos.total_held),
        ("Revisadas y etiquetadas", datos.total_reviewed),
        ("Fraudes confirmados", datos.total_actual_frauds),
        ("Fraudes detectados", datos.total_detected_frauds),
        ("Fraudes no detectados", datos.total_undetected_frauds),
    ]
    for i, (etiqueta, valor) in enumerate(base):
        fila = 15 + i
        hoja.cell(row=fila, column=1, value=etiqueta).font = Font(size=10)
        hoja.cell(row=fila, column=2, value=valor).font = Font(bold=True, size=10)

    # La advertencia va en el archivo y no solo en la pantalla, porque el
    # archivo es lo que acaba pegado en el documento de la tesis.
    fila = 15 + len(base) + 1
    hoja.cell(
        row=fila,
        column=1,
        value=(
            "Nota: las dos tasas se calculan solo sobre los pedidos que un administrador "
            "revisó y etiquetó. Un pedido bloqueado nunca llega a cobrarse, así que nunca "
            "tendrá un contracargo que lo confirme como fraude: los aciertos más valiosos "
            "del modelo son también los más difíciles de etiquetar, y estas cifras los "
            "subestiman."
        ),
    ).alignment = Alignment(wrap_text=True, vertical="top")
    hoja.merge_cells(start_row=fila, start_column=1, end_row=fila + 2, end_column=4)


def _hoja_serie(libro: Workbook, datos: FraudHistoryResponse) -> None:
    hoja = libro.create_sheet("Indicadores por período")
    _anchos(hoja, [14, 12, 11, 12, 12, 11, 11, 12, 13, 14, 14, 15, 15])

    _encabezado(
        hoja,
        1,
        [
            "Período",
            "Evaluadas",
            "Pasaron",
            "En revisión",
            "Bloqueadas",
            "Revisadas",
            "Fraudes",
            "Detectados",
            "No detectados",
            "Tasa detectados",
            "Tasa no detectado",
            "Tiempo medio (ms)",
            "Puntaje medio",
        ],
    )
    hoja.row_dimensions[1].height = 30
    hoja.freeze_panes = "B2"

    for i, p in enumerate(datos.periods):
        fila = 2 + i
        # La fecha va como fecha de verdad y no como texto: así Excel puede
        # ordenarla y filtrarla, que es la mitad de la razón para exportar.
        celda = hoja.cell(row=fila, column=1, value=p.period_start)
        celda.number_format = "DD/MM/YYYY"

        hoja.cell(row=fila, column=2, value=p.evaluations)
        hoja.cell(row=fila, column=3, value=p.approved)
        hoja.cell(row=fila, column=4, value=p.in_review)
        hoja.cell(row=fila, column=5, value=p.blocked)
        hoja.cell(row=fila, column=6, value=p.reviewed)
        hoja.cell(row=fila, column=7, value=p.actual_frauds)
        hoja.cell(row=fila, column=8, value=p.detected_frauds)
        hoja.cell(row=fila, column=9, value=p.undetected_frauds)

        # Las tasas van como número con formato de porcentaje, no como el texto
        # "57,1 %": un texto no se promedia ni se grafica.
        for columna, tasa in ((10, p.detection_rate), (11, p.undetected_rate)):
            celda = hoja.cell(row=fila, column=columna, value=tasa)
            celda.number_format = "0.0%"
            if tasa is None:
                celda.value = "sin datos"
                celda.font = Font(color="9AA5B1", italic=True)

        hoja.cell(row=fila, column=12, value=p.average_detection_time_ms).number_format = "0.0"
        hoja.cell(row=fila, column=13, value=p.average_score).number_format = "0.000"

    hoja.auto_filter.ref = f"A1:M{1 + len(datos.periods)}"


def _hoja_montos(libro: Workbook, datos: FraudHistoryResponse) -> None:
    hoja = libro.create_sheet("Montos")
    _anchos(hoja, [16, 20, 20, 20])

    _encabezado(hoja, 1, ["Período", "Monto cobrable", "Monto retenido", "Total evaluado"])
    hoja.row_dimensions[1].height = 26

    for i, p in enumerate(datos.periods):
        fila = 2 + i
        hoja.cell(row=fila, column=1, value=p.period_start).number_format = "DD/MM/YYYY"
        for columna, monto in (
            (2, p.approved_amount),
            (3, p.held_amount),
            (4, p.approved_amount + p.held_amount),
        ):
            hoja.cell(row=fila, column=columna, value=monto).number_format = '"S/" #,##0.00'

    total = 2 + len(datos.periods)
    hoja.cell(row=total, column=1, value="Total").font = Font(bold=True)
    for columna in (2, 3, 4):
        letra = get_column_letter(columna)
        # Una fórmula y no el número ya sumado: quien abra el archivo puede
        # filtrar filas y ver el total recalcularse.
        celda = hoja.cell(row=total, column=columna, value=f"=SUM({letra}2:{letra}{total - 1})")
        celda.number_format = '"S/" #,##0.00'
        celda.font = Font(bold=True)


def _desde_la_primera_actividad(datos: FraudHistoryResponse) -> FraudHistoryResponse:
    """
    Recorta los períodos vacíos que anteceden al primer movimiento.

    La pantalla los dibuja a propósito —en una gráfica, los días tranquilos son
    parte de la forma de la curva—, pero en una hoja de cálculo son diecinueve
    filas de ceros antes del primer dato, y quien abre el archivo se encuentra
    con eso. Los huecos interiores sí se conservan: un mes sin ventas entre dos
    con ventas es información.

    Si no hubo actividad en toda la ventana no se recorta nada, para que el
    archivo no salga vacío y sin explicación.
    """
    primero = next(
        (i for i, p in enumerate(datos.periods) if p.evaluations > 0), None
    )
    if primero is None or primero == 0:
        return datos
    return datos.model_copy(update={"periods": datos.periods[primero:]})


def construir(datos: FraudHistoryResponse) -> BytesIO:
    """Arma el libro y lo devuelve listo para enviar."""
    datos = _desde_la_primera_actividad(datos)
    libro = Workbook()
    _hoja_portada(libro, datos)
    _hoja_serie(libro, datos)
    _hoja_montos(libro, datos)

    memoria = BytesIO()
    libro.save(memoria)
    memoria.seek(0)
    return memoria


__all__ = ["construir", "INDICADORES", "ESCALAS"]
